"""
excel_parser.py — Lecture du tableau Excel "Formations et inscrits"

Structure du fichier Excel :
  Ligne 1  : Nom client (B-D mergées) + Type de groupe par colonne (Administrateurs, Utilisateurs…)
  Ligne 2  : Domaine du module (Socle commun, CRM, SAV, Facturation…)
  Ligne 3  : Durée en demi-journées (toujours 0.5) — discriminant des colonnes modules
  Ligne 4  : Date prévisionnelle (peut être vide avant planification)
  Ligne 5  : Horaires (peut être vide avant planification)
  Ligne 6  : Modalités (VISIO ou SUR SITE)
  Ligne 7  : Capacité min/max
  Ligne 8  : Vide (séparateur)
  Lignes 9+: Stagiaires (N°, NOM, Prénom, Email, puis marks par colonne module)

Marks possibles dans les colonnes modules :
  "x"          → participant confirmé
  "TNS"        → travailleur non salarié (participe mais exclu de la convention)
  "optionnel"  → à inclure par défaut, Estelle peut décocher
  "à inviter"  → idem
  None / ""    → ne participe pas à ce module
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl


# ---------------------------------------------------------------------------
# Types de participation
# ---------------------------------------------------------------------------

MARK_CONFIRMED = "x"
MARK_TNS = "TNS"
MARK_OPTIONAL = "optionnel"
MARK_TO_INVITE = "à inviter"
ACTIVE_MARKS = {MARK_CONFIRMED, MARK_TNS, MARK_OPTIONAL, MARK_TO_INVITE}

# Groupes qui ne sont PAS des modules de formation à facturer
NON_FORMATION_GROUPS = {"Interne Koban"}

# Lignes de métadonnées (1-indexé, correspondant aux lignes Excel)
ROW_TYPE_GROUPE = 1
ROW_DOMAINE = 2
ROW_DUREE = 3
ROW_DATE = 4
ROW_HORAIRES = 5
ROW_MODALITES = 6
ROW_CAPACITE = 7

# Colonnes fixes des stagiaires (1-indexé)
COL_NUM = 1
COL_NOM = 2
COL_PRENOM = 3
COL_EMAIL = 4
FIRST_MODULE_COL = 5  # Première colonne de module (col E)

ROW_FIRST_STAGIAIRE = 9  # Les stagiaires commencent ligne 9


# ---------------------------------------------------------------------------
# Dataclasses de sortie
# ---------------------------------------------------------------------------

@dataclass
class ModuleColumn:
    """Métadonnées d'une colonne module dans le tableau Excel."""
    col_index: int          # Index 1-based dans Excel (5 = col E)
    type_groupe: str        # "Administrateurs" | "Utilisateurs" | …
    domaine: str            # "Socle commun" | "CRM" | "SAV" | …
    duree: float            # Toujours 0.5 (demi-journée)
    date: Optional[str]     # Date prévisionnelle (None si non remplie)
    horaires: Optional[str] # Horaires (None si non remplis)
    modalite: str           # "VISIO" | "SUR SITE"
    capacite: Optional[str] # Texte capacité min/max


@dataclass
class Stagiaire:
    """Un stagiaire avec ses participations par colonne module."""
    num: int
    nom: str
    prenom: str
    email: str
    # col_index (1-based) → mark ("x", "TNS", "optionnel", "à inviter")
    modules: dict[int, str] = field(default_factory=dict)

    @property
    def nom_complet(self) -> str:
        return f"{self.nom} {self.prenom}"

    def is_tns(self) -> bool:
        """Vrai si tous les marks sont TNS (travailleur non salarié)."""
        return all(m == MARK_TNS for m in self.modules.values())

    def active_col_indexes(self, include_optional: bool = True) -> list[int]:
        """Retourne les col_index où le stagiaire est actif."""
        result = []
        for col_idx, mark in self.modules.items():
            if mark == MARK_CONFIRMED:
                result.append(col_idx)
            elif mark in (MARK_OPTIONAL, MARK_TO_INVITE) and include_optional:
                result.append(col_idx)
            # TNS : jamais inclus dans active_col_indexes
        return sorted(result)


@dataclass
class ExcelData:
    """Résultat complet du parsing du fichier Excel."""
    client_name: str
    module_columns: list[ModuleColumn]
    stagiaires: list[Stagiaire]

    def module_col_by_index(self, col_index: int) -> Optional[ModuleColumn]:
        for mc in self.module_columns:
            if mc.col_index == col_index:
                return mc
        return None


# ---------------------------------------------------------------------------
# Parser principal
# ---------------------------------------------------------------------------

def parse_formations_excel(file_path: str | Path) -> ExcelData:
    """
    Parse le fichier Excel "Formations et inscrits".

    Args:
        file_path: Chemin vers le .xlsx

    Returns:
        ExcelData avec les colonnes modules et la liste des stagiaires.

    Raises:
        ValueError: Si l'onglet attendu n'est pas trouvé ou structure invalide.
        FileNotFoundError: Si le fichier n'existe pas.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Fichier Excel introuvable : {path}")

    wb = openpyxl.load_workbook(path, data_only=True)

    # Sélection de l'onglet
    ws = _select_worksheet(wb)

    # Lecture brute de toutes les cellules en une passe (dict row -> col -> value)
    raw = _read_all_cells(ws)

    # Nom du client (ligne 1, col B — cellules B1:D1 fusionnées)
    client_name = str(raw.get(ROW_TYPE_GROUPE, {}).get(COL_NOM, "")).strip() or "CLIENT"

    # Identification des colonnes modules (col E et suivantes avec durée = 0.5)
    module_columns = _extract_module_columns(raw, ws.max_column)

    # Lecture des stagiaires (lignes 9+)
    stagiaires = _extract_stagiaires(raw, ws.max_row, module_columns)

    return ExcelData(
        client_name=client_name,
        module_columns=module_columns,
        stagiaires=stagiaires,
    )


# ---------------------------------------------------------------------------
# Fonctions internes
# ---------------------------------------------------------------------------

def _select_worksheet(wb: openpyxl.Workbook):
    """Sélectionne l'onglet 'Formations et inscrits', sinon le premier."""
    target = "Formations et inscrits"
    if target in wb.sheetnames:
        return wb[target]
    return wb.active


def _read_all_cells(ws) -> dict[int, dict[int, object]]:
    """
    Lit toutes les cellules en un dict {row: {col: value}}.
    Ignore les cellules vides (None ou chaîne vide).
    Gère les cellules fusionnées en propageant la valeur de la cellule top-left.
    """
    # Reconstruire les valeurs des cellules fusionnées
    merged_values: dict[tuple[int, int], object] = {}
    for merged_range in ws.merged_cells.ranges:
        top_left = ws.cell(merged_range.min_row, merged_range.min_col)
        val = top_left.value
        if val is None:
            continue
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, col)] = val

    result: dict[int, dict[int, object]] = {}
    for row in ws.iter_rows():
        for cell in row:
            val = merged_values.get((cell.row, cell.column), cell.value)
            if val is None or (isinstance(val, str) and val.strip() == ""):
                continue
            result.setdefault(cell.row, {})[cell.column] = val

    return result


def _extract_module_columns(raw: dict, max_col: int) -> list[ModuleColumn]:
    """
    Identifie les colonnes de modules de formation.
    Critère principal : la ligne 3 (ROW_DUREE) contient une valeur numérique (0.5).
    Exclusion : colonnes dont le type_groupe est dans NON_FORMATION_GROUPS.
    """
    row1 = raw.get(ROW_TYPE_GROUPE, {})
    row2 = raw.get(ROW_DOMAINE, {})
    row3 = raw.get(ROW_DUREE, {})
    row4 = raw.get(ROW_DATE, {})
    row5 = raw.get(ROW_HORAIRES, {})
    row6 = raw.get(ROW_MODALITES, {})
    row7 = raw.get(ROW_CAPACITE, {})

    columns = []
    for col in range(FIRST_MODULE_COL, max_col + 1):
        duree_val = row3.get(col)

        # Colonne module = durée numérique présente
        if duree_val is None or not isinstance(duree_val, (int, float)):
            continue

        type_groupe = str(row1.get(col, "")).strip()
        domaine = str(row2.get(col, "")).strip()

        # Exclusion des groupes non-formation
        if type_groupe in NON_FORMATION_GROUPS:
            continue

        date_val = row4.get(col)
        horaires_val = row5.get(col)
        modalite_val = str(row6.get(col, "")).strip()
        capacite_val = row7.get(col)

        columns.append(ModuleColumn(
            col_index=col,
            type_groupe=type_groupe,
            domaine=domaine,
            duree=float(duree_val),
            date=str(date_val).strip() if date_val is not None else None,
            horaires=str(horaires_val).strip() if horaires_val is not None else None,
            modalite=modalite_val,
            capacite=str(capacite_val).strip() if capacite_val is not None else None,
        ))

    return columns


def _extract_stagiaires(
    raw: dict,
    max_row: int,
    module_columns: list[ModuleColumn],
) -> list[Stagiaire]:
    """
    Lit les stagiaires à partir de la ligne ROW_FIRST_STAGIAIRE.
    S'arrête à la première ligne sans NOM (fin du tableau).
    """
    module_col_indexes = {mc.col_index for mc in module_columns}
    stagiaires = []

    for row_num in range(ROW_FIRST_STAGIAIRE, max_row + 1):
        row_data = raw.get(row_num, {})

        nom = str(row_data.get(COL_NOM, "")).strip()
        if not nom:
            continue  # Ligne vide ou hors tableau

        num_val = row_data.get(COL_NUM)
        num = int(num_val) if isinstance(num_val, (int, float)) else 0
        prenom = str(row_data.get(COL_PRENOM, "")).strip()
        email = str(row_data.get(COL_EMAIL, "")).strip()

        # Lecture des marks pour chaque colonne module
        modules: dict[int, str] = {}
        for col_idx in module_col_indexes:
            raw_mark = row_data.get(col_idx)
            if raw_mark is None:
                continue
            mark = str(raw_mark).strip().lower()
            # Normalisation des marks
            if mark == "x":
                modules[col_idx] = MARK_CONFIRMED
            elif mark == "tns":
                modules[col_idx] = MARK_TNS
            elif mark in ("optionnel", "optional"):
                modules[col_idx] = MARK_OPTIONAL
            elif mark in ("à inviter", "a inviter", "à inviter"):
                modules[col_idx] = MARK_TO_INVITE

        if not modules and num == 0:
            continue  # Ignorer les lignes sans données utiles

        stagiaires.append(Stagiaire(
            num=num,
            nom=nom,
            prenom=prenom,
            email=email,
            modules=modules,
        ))

    return stagiaires
